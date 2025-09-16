from __future__ import annotations

from datetime import datetime, date
from typing import Optional
from pathlib import Path
import os
import tempfile
import unicodedata
import re

from openpyxl import load_workbook, Workbook
try:
    # Utilidad para convertir números de serie de Excel a datetime/date
    from openpyxl.utils.datetime import from_excel as _from_excel
except Exception:  # pragma: no cover - fallback si cambia API
    _from_excel = None

from .db import TMPActividades, get_session_factory


def _parse_horas(hours_value, minutes_value) -> str:
    try:
        hours = int(hours_value or 0)
    except Exception:
        hours = 0
    try:
        minutes = int(minutes_value or 0)
    except Exception:
        minutes = 0
    # Normalizar minutos > 59
    total_minutes = hours * 60 + minutes
    hh = total_minutes // 60
    mm = total_minutes % 60
    return f"{hh:02d}:{mm:02d}:00"


def _parse_fecha(value, epoch: str = "windows") -> Optional[date]:
    if value is None or value == "":
        return None
    if isinstance(value, (datetime, date)):
        return value.date() if isinstance(value, datetime) else value
    # Si llega como número de serie Excel (float/int), intentar convertir
    if isinstance(value, (int, float)) and _from_excel is not None:
        try:
            dt = _from_excel(value, epoch=epoch)
            return dt.date()
        except Exception:
            pass
    # Intentar varios formatos comunes
    for fmt in ("%d/%m/%Y", "%Y-%m-%d", "%m/%d/%Y"):
        try:
            return datetime.strptime(str(value), fmt).date()
        except Exception:
            continue
    # Fallback: no convertir
    return None


def _norm(text) -> str:
    """Normaliza un texto: minúsculas, sin acentos, espacios colapsados."""
    if text is None:
        return ""
    s = str(text).strip().lower()
    s = "".join(c for c in unicodedata.normalize("NFD", s) if unicodedata.category(c) != "Mn")
    return " ".join(s.split())


def _parse_numero_act(value) -> Optional[int]:
    """Extrae el último segmento numérico del campo 'Número' y lo convierte a entero.

    Formato esperado: 'dddd-xxx.xxx' donde dddd son 4 dígitos y xxx.xxx son 6 dígitos con punto de miles.
    Se extrae el segmento derecho, se eliminan separadores y se convierte a int.
    """
    if value is None:
        return None
    s = str(value).strip()
    if s == "":
        return None
    # Tomar el segmento a la derecha del guión si existe
    if "-" in s:
        right = s.split("-")[-1].strip()
    else:
        right = s
    # Mantener solo dígitos
    digits_only = "".join(ch for ch in right if ch.isdigit())
    if digits_only == "":
        return None
    try:
        return int(digits_only)
    except Exception:
        return None


def _find_headers(ws, candidate_rows=None) -> tuple[dict[str, int], int]:
    """Busca encabezados en filas candidatas y devuelve (mapa_normalizado, fila).

    Si no se especifican filas, escanea automáticamente desde la 1 hasta la 20 (o max_row).
    """
    if candidate_rows is None:
        upper = min(20, ws.max_row or 20)
        candidate_rows = range(1, upper + 1)
    for r in candidate_rows:
        headers: dict[str, int] = {}
        row = ws[r]
        if not row:
            continue
        for idx, cell in enumerate(row, start=1):
            key = _norm(cell.value) or f"col{idx}"
            headers[key] = idx
        if any(k in headers for k in ("fecha", "numero", "número", "asunto")):
            return headers, r
    return {}, -1


def _is_empty(value) -> bool:
    """True si el valor es None o string vacío (tras strip)."""
    if value is None:
        return True
    if isinstance(value, str):
        return value.strip() == ""
    return False


def _is_header_like(fecha_val, numero_val, asunto_val) -> bool:
    """Detecta si una fila parece repetir encabezados (p. ej., 'Fecha', 'Número', 'Asunto')."""
    values = {_norm(fecha_val), _norm(numero_val), _norm(asunto_val)}
    targets = {"fecha", "numero", "número", "asunto"}
    return len(values & targets) >= 1


def _import_xlsx(file_path: str) -> int:
    """
    Importa actividades desde un archivo Excel a la tabla TMP_Actividades.

    Reglas específicas:
    - A3: Numero_Responsable (entero) a repetir en cada registro
    - Comenzar desde fila 4
    - Columnas:
        Fecha -> col 'Fecha'
        Numero_Act -> col 'Número'
        Asunto -> col 'Asunto'
        Horas -> combinar col J (horas) y col L (minutos) en HH:MM:SS
    """
    wb = load_workbook(filename=file_path, data_only=True)
    ws = wb.active

    # Numero Responsable en A3
    num_resp_cell = ws["A3"].value
    try:
        # Extraer el número de responsable desde el formato 'Responsable:     195     Dario'
        import re
        match = re.search(r"(\d+)", str(num_resp_cell))
        numero_responsable = int(match.group(1)) if match else 0

        # Extraer el nombre de responsable, a continuación del número, desde el formato 'Responsable:     195     Dario'
        match = re.search(r"(\w+)", str(num_resp_cell))
        nombre_responsable = match.group(1) if match else ""
    except Exception:
        numero_responsable = 0

    # Identificar encabezados por nombre para 'Fecha', 'Número', 'Asunto'
    headers, header_row_index = _find_headers(ws)
    if header_row_index == -1:
        raise RuntimeError("No se encontraron encabezados en filas 3 o 4.")

    def get_any(options: list[str]) -> int | None:
        for opt in options:
            key = _norm(opt)
            if key in headers:
                return headers[key]
        return None

    col_fecha = get_any(["fecha"])  
    col_numero = get_any(["numero", "número", "numero act", "num act", "nro", "nº", "no"])  
    col_asunto = get_any(["asunto", "asuntos", "descripcion", "descripción", "concepto", "detalle", "subject"])  

    missing = [name for name, idx in (("Fecha", col_fecha), ("Número", col_numero), ("Asunto", col_asunto)) if not idx]
    if missing:
        found = ", ".join(headers.keys())
        raise RuntimeError(f"Faltan columnas requeridas: {', '.join(missing)}. Encabezados detectados: {found}")

    # Las columnas de horas/minutos son fijas según requerimiento: J y L
    col_horas_idx = 10  # J
    col_minutos_idx = 12  # L

    # Detectar epoch del workbook para convertir seriales Excel correctamente
    epoch_flag = "windows"
    try:
        epoch_dt = getattr(wb, "epoch", None)
        if epoch_dt is not None and getattr(epoch_dt, "year", 1900) == 1904:
            epoch_flag = "mac"
    except Exception:
        pass

    SessionFactory = get_session_factory()
    session = SessionFactory()
    try:
        # Limpiar tabla
        session.query(TMPActividades).delete()
        session.commit()

        inserted = 0
        start_row = header_row_index + 1
        # Garantizar que nunca se lea antes de la fila 4 (A3 contiene solo responsable)
        if start_row < 4:
            start_row = 4
        # Si la fila inmediatamente posterior a encabezados está vacía o parece otro encabezado, saltarla
        try:
            first_fecha = ws.cell(row=start_row, column=col_fecha).value if col_fecha else None
            first_numero = ws.cell(row=start_row, column=col_numero).value if col_numero else None
            first_asunto = ws.cell(row=start_row, column=col_asunto).value if col_asunto else None
            if (_is_empty(first_fecha) and _is_empty(first_numero) and _is_empty(first_asunto)) or _is_header_like(first_fecha, first_numero, first_asunto):
                start_row += 1
        except Exception:
            pass
        for row_idx in range(start_row, ws.max_row + 1):
            # Condición de finalización: si la columna A contiene un valor no-fecha, terminar importación
            try:
                col_a_val = ws.cell(row=row_idx, column=1).value
                if not _is_empty(col_a_val):
                    if _parse_fecha(col_a_val, epoch=epoch_flag) is None:
                        break
            except Exception:
                # Si ocurre algún error inesperado al leer/parsear, continuar con la lógica estándar
                pass
            fecha_val = ws.cell(row=row_idx, column=col_fecha).value if col_fecha else None
            numero_val = ws.cell(row=row_idx, column=col_numero).value if col_numero else None
            asunto_val = ws.cell(row=row_idx, column=col_asunto).value if col_asunto else None
            horas_val = ws.cell(row=row_idx, column=col_horas_idx).value
            minutos_val = ws.cell(row=row_idx, column=col_minutos_idx).value

            # Fila vacía aparente (previa al parse) o fila con texto de encabezado
            if _is_empty(fecha_val) and _is_empty(numero_val) and _is_empty(asunto_val):
                continue
            if _is_header_like(fecha_val, numero_val, asunto_val):
                continue

            fecha = _parse_fecha(fecha_val, epoch=epoch_flag)
            numero_act = None if _is_empty(numero_val) else _parse_numero_act(numero_val)
            asunto = None if _is_empty(asunto_val) else str(asunto_val).strip()

            # Fila vacía real tras normalización/parse: no insertar
            if fecha is None and numero_act is None and asunto is None:
                continue
            horas = _parse_horas(horas_val, minutos_val)

            session.add(
                TMPActividades(
                    numero_responsable=numero_responsable,
                    fecha=fecha,
                    numero_act=numero_act,
                    asunto=asunto,
                    horas=horas,
                )
            )
            inserted += 1

        session.commit()
        return inserted
    finally:
        session.close()


def _convert_xls_to_temp_xlsx(file_path: str) -> str:
    """
    Convierte un archivo .xls a un .xlsx temporal copiando únicamente valores
    (y convirtiendo fechas correctamente) para ser procesado con openpyxl.

    Retorna la ruta del archivo .xlsx temporal generado. El llamador es
    responsable de eliminarlo.
    """
    try:
        import xlrd
    except Exception as e:
        raise RuntimeError("El formato .xls requiere la librería 'xlrd'.") from e

    book = xlrd.open_workbook(file_path)
    sheet = book.sheet_by_index(0)

    # Crear archivo temporal .xlsx (cerrar descriptor para Windows)
    fd, temp_path = tempfile.mkstemp(suffix=".xlsx")
    os.close(fd)

    wb_out = Workbook()
    ws_out = wb_out.active

    for r in range(sheet.nrows):
        for c in range(sheet.ncols):
            cell_type = sheet.cell_type(r, c)
            cell_value = sheet.cell_value(r, c)

            # Convertir fechas desde número serial Excel
            if cell_type == xlrd.XL_CELL_DATE:
                try:
                    y, m, d, hh, mm, ss = xlrd.xldate_as_tuple(cell_value, book.datemode)
                    # Si hay hora 0,0,0 lo escribimos como date; si no, datetime
                    if (hh, mm, ss) == (0, 0, 0):
                        value_converted = date(y, m, d)
                    else:
                        value_converted = datetime(y, m, d, hh, mm, ss)
                except Exception:
                    value_converted = cell_value
            else:
                value_converted = cell_value

            ws_out.cell(row=r + 1, column=c + 1, value=value_converted)

    wb_out.save(temp_path)
    return temp_path


def import_actividades_from_excel(file_path: str) -> int:
    ext = Path(file_path).suffix.lower()
    if ext == ".xlsx":
        return _import_xlsx(file_path)
    if ext == ".xls":
        temp_xlsx: Optional[str] = None
        try:
            temp_xlsx = _convert_xls_to_temp_xlsx(file_path)
            return _import_xlsx(temp_xlsx)
        finally:
            if temp_xlsx and os.path.exists(temp_xlsx):
                try:
                    os.remove(temp_xlsx)
                except Exception:
                    pass
    raise RuntimeError("Extensión no soportada. Usa .xls o .xlsx")


