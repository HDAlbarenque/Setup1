from __future__ import annotations

from datetime import datetime
from typing import Optional
from pathlib import Path
import os
import tempfile

from openpyxl import load_workbook, Workbook

try:
    from openpyxl.utils.datetime import from_excel as _from_excel
except Exception:
    _from_excel = None

from .db import TMPActividadesDario, get_session_factory


def _parse_datetime(value, epoch: str = "windows") -> Optional[datetime]:
    """Parsea un valor a datetime, manejando seriales de Excel y strings."""
    if value is None or value == "":
        return None
    if isinstance(value, datetime):
        return value
    if isinstance(value, (int, float)) and _from_excel is not None:
        try:
            return _from_excel(value, epoch=epoch)
        except Exception:
            pass
    if isinstance(value, str):
        for fmt in ("%Y-%m-%d %H:%M:%S", "%d/%m/%Y %H:%M", "%d/%m/%Y"):
            try:
                return datetime.strptime(value, fmt)
            except Exception:
                continue
    return None


def _is_empty(value) -> bool:
    """True si el valor es None o string vacío (tras strip)."""
    if value is None:
        return True
    if isinstance(value, str):
        return value.strip() == ""
    return False


def _import_xlsx_dario(file_path: str, numero_responsable: int) -> int:
    """Importa datos desde un archivo Excel a la tabla TMP_Actividades_Dario."""
    wb = load_workbook(filename=file_path, data_only=True)
    ws = wb.active
    
    epoch_flag = "windows"
    try:
        if getattr(getattr(wb, "epoch", None), "year", 1900) == 1904:
            epoch_flag = "mac"
    except Exception:
        pass

    SessionFactory = get_session_factory()
    session = SessionFactory()
    try:
        # Limpiar tabla antes de importar
        session.query(TMPActividadesDario).delete()
        session.commit()

        inserted = 0
        # Omitir la primera fila (encabezados)
        for row_idx in range(2, ws.max_row + 1):
            row_values = [ws.cell(row=row_idx, column=c).value for c in range(1, 12)]
            
            # Si las primeras columnas importantes están vacías, saltar fila
            if all(_is_empty(v) for v in row_values[:3]):
                continue

            # Mapeo de columnas a modelo
            actividad = TMPActividadesDario(
                size=str(row_values[0]) if not _is_empty(row_values[0]) else None,
                numero=int(row_values[1]) if not _is_empty(row_values[1]) else None,
                nombre=str(row_values[2]) if not _is_empty(row_values[2]) else None,
                comienzo=_parse_datetime(row_values[3], epoch=epoch_flag),
                fin=_parse_datetime(row_values[4], epoch=epoch_flag),
                sintesis=str(row_values[5]) if not _is_empty(row_values[5]) else None,
                observaciones=str(row_values[6]) if not _is_empty(row_values[6]) else None,
                vcx_s=str(row_values[7]) if not _is_empty(row_values[7]) else None,
                req_sincro=str(row_values[8]) if not _is_empty(row_values[8]) else None,
                version=str(row_values[9]) if not _is_empty(row_values[9]) else None,
                numero_responsable=numero_responsable,
            )
            session.add(actividad)
            inserted += 1

        session.commit()
        return inserted
    finally:
        session.close()


def _convert_xls_to_temp_xlsx(file_path: str) -> str:
    """Convierte un archivo .xls a un .xlsx temporal."""
    try:
        import xlrd
    except Exception as e:
        raise RuntimeError("El formato .xls requiere la librería 'xlrd'.") from e

    book = xlrd.open_workbook(file_path)
    sheet = book.sheet_by_index(0)

    fd, temp_path = tempfile.mkstemp(suffix=".xlsx")
    os.close(fd)

    wb_out = Workbook()
    ws_out = wb_out.active

    for r in range(sheet.nrows):
        for c in range(sheet.ncols):
            cell_type = sheet.cell_type(r, c)
            cell_value = sheet.cell_value(r, c)
            if cell_type == xlrd.XL_CELL_DATE:
                try:
                    dt_tuple = xlrd.xldate_as_tuple(cell_value, book.datemode)
                    value_converted = datetime(*dt_tuple)
                except Exception:
                    value_converted = cell_value
            else:
                value_converted = cell_value
            ws_out.cell(row=r + 1, column=c + 1, value=value_converted)

    wb_out.save(temp_path)
    return temp_path


def import_actividades_dario(file_path: str, numero_responsable: int) -> int:
    """Función principal para importar actividades desde un archivo Excel."""
    ext = Path(file_path).suffix.lower()
    if ext == ".xlsx":
        return _import_xlsx_dario(file_path, numero_responsable)
    if ext == ".xls":
        temp_xlsx: Optional[str] = None
        try:
            temp_xlsx = _convert_xls_to_temp_xlsx(file_path)
            return _import_xlsx_dario(temp_xlsx, numero_responsable)
        finally:
            if temp_xlsx and os.path.exists(temp_xlsx):
                try:
                    os.remove(temp_xlsx)
                except Exception:
                    pass
    raise RuntimeError("Extensión no soportada. Usa .xls o .xlsx")

